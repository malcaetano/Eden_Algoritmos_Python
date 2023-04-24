#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#     Programa para buscar  palavras-chave no Excel
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
from textblob import TextBlob as tb
import tweepy
import numpy as np
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as fig
from unidecode import unidecode
from openpyxl import load_workbook
from deep_translator import GoogleTranslator
import re
from wordcloud import WordCloud,STOPWORDS
import matplotlib.pyplot as fig

consumer_key = 'MEbEz5ewLzf0lOqg74AdY0GHB'
consumer_secret = 'K2QziIjFVVrl9omjAkno4HXyeC8uEDvvhw4WH9ZOzWKwFMo3QJ'
access_token = '149650363-30wY3uNi340toTtqzO8gFaV5RGL7khOn07t6Q9kh'
access_token_secret = 'zH9RvXiSuQSN5Igrenw11Liyh1mpsekDl0mb5No8Lsv4D'


auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token, access_token_secret)
api = tweepy.API(auth)

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
dados=pd.read_excel('Ent_Pol.xlsx',sheet_name='Planilha1')
df=pd.DataFrame(dados)
palchv=0
x=np.zeros(21)
for j in range(4):
    klist=dados[dados.columns[j]].tolist()
    for i in range(len(klist)):  
           public_tweets1 = api.search(klist[i],count=5,result_type='recent')
           analysis = None
           tweets1 = [] # Lista vazia para armazenar scores
           for tweet in public_tweets1:
               try:
                       textPort=unidecode(tweet.text)
                       processed_text = re.sub(r"(!?:\@|http?\://|https?\://|www)\S+", "", tweet.text)
                       processed_text = re.sub(r"^RT[\s]+,", "",processed_text)
                       processed_text = " ".join(processed_text.split())
                       processed_text = processed_text.lower()
                       textIng=GoogleTranslator(source='auto',target='english').translate(processed_text)
                       analysis = tb(textIng)
                       polarity1 = analysis.sentiment.polarity
                       tweets1.append(polarity1)
               except:
                   pass
           x[palchv]=np.mean(tweets1)
           x[np.isnan(x)]=0
           print('++++++++++++++++++++++++++++++++++++++++++++')  
           print('PALAVRA: ' + str(klist[i]))                     
           print('MÉDIA DE SENTIMENTO: ' + str(x[palchv]))
           palchv=palchv+1
           
med_geral=np.mean(x)
if med_geral<0:
    status='negativo'
else:
    status='positivo-neutro'
    
print('++++++++++++++++++++++++++++++++++++++++++++')                       
print('MÉDIA GERAL DOS SENTIMENTOS: ' + str(med_geral))
print('STATUS DO SENTIMENTO: ', status)

#+++++++++++++++++++++++++++ salvando os dados no arquivo texto +++++++++++
#+++++++++++++++++++++++++++ salvando os dados no arquivo texto +++++++++++
data=datetime.now()
data=data.strftime("%Y-%m-%d %H:%M:%S")
sensacao1=status
sensacao2=med_geral
sent1=x[0]  # hospital
sent2=x[4]  # tosse
sent3=x[1]  # pressao alta
sent4=x[2]  # dengue
sent5=x[3]  # sarampo
sent6=x[5]  # tuberculose
sent7=x[6]  # garganta inflamada
sent8=x[7]  # figado
sent9=x[8]  # infeccao antibiotico
sent10=x[9] # enxaqueca
sent11=x[10] # pneumonia
sent12=x[11] # alergia
sent13=x[12] # virose
sent14=x[13] # diarreia
sent15=x[14] # covid
sent16=x[15] # resfriado
sent17=x[16] # febre
sent18=x[17] # coceira
sent19=x[18] # doar sangue
sent20=x[19] # asma
df=pd.DataFrame({'A':[data],
                     'B':[sensacao1],
                     'C':[sensacao2],
                     'D':[sent1],
                     'E':[sent2],
                     'F':[sent3],
                     'G':[sent4],
                     'H':[sent5],
                     'I':[sent6],
                     'J':[sent7],
                     'K':[sent8],
                     'L':[sent9],
                     'M':[sent10],
                     'N':[sent11],
                     'O':[sent12],
                     'P':[sent13],
                     'Q':[sent14],
                     'R':[sent15],
                     'S':[sent16],
                     'T':[sent17],
                     'U':[sent18],
                     'V':[sent19],
                     'W':[sent20],})
book = load_workbook('Res_pol.xlsx')
writer = pd.ExcelWriter('Res_Pol.xlsx',engine='openpyxl',mode='a')
writer.book=book
writer.sheets = {ws.title: ws for ws in book.worksheets}
df.to_excel(writer,sheet_name='Planilha1',startrow = writer.sheets['Planilha1'].max_row,
            index=False,
            header=False)
writer.save()
writer.close()

#++++++++++++++++++++++++ abrindo os dados para fazer o gráfico +++++++++++
dados=pd.read_excel('Res_Pol.xlsx',sheet_name='Planilha1')
df=pd.DataFrame(dados)

ax1=fig.subplot(111)
ax1.plot(df['data'],df['valor'],'-k',linewidth=3)
ax1.fill_between(df['data'],df['valor'], where=df['valor'] >= 0, 
                 facecolor='blue', interpolate=True,alpha=0.4)
ax1.fill_between(df['data'],df['valor'], where=df['valor'] < 0, 
                 facecolor='red', interpolate=True,alpha=0.4)
ax1.set_ylabel('Sentimento',fontsize=18,weight='bold')
fig.gcf().autofmt_xdate()
ax1.grid()

#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++




